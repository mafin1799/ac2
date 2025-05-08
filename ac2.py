import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from scipy.stats import skew, kurtosis
from typing import List, Tuple, Optional

# Константы
FILE_PATH = 'AsOr_Zadanie_1_2.xlsx'
SHEET_NAMES = {
    'uniform': 'Равномерное',
    'normal': 'Нормальное',
    'gamma': 'Гамма',
    'beta': 'Бета'
}


def load_workbook():
    """Загрузка Excel-файла"""
    return openpyxl.load_workbook(FILE_PATH, data_only=True)


def read_distribution_data(workbook, dist_type: str) -> Tuple:
    """Чтение параметров распределения из Excel"""
    sheet = workbook[SHEET_NAMES[dist_type]]

    params = {
        'uniform': lambda: (
            float(sheet['A5'].value),  # a
            float(sheet['B5'].value),  # b
            int(sheet['B6'].value),  # k
            int(sheet['F11'].value),  # s
            read_sample(sheet, 'B9', int(sheet['B6'].value))  # sample
        ),
        'normal': lambda: (
            float(sheet['B4'].value),  # mean
            float(sheet['B5'].value),  # std_dev
            int(sheet['B6'].value),  # k
            int(sheet['G11'].value),  # s
            read_sample(sheet, 'B9', int(sheet['B6'].value))  # sample
        ),
        'gamma': lambda: (
            float(sheet['B4'].value),  # alpha
            float(sheet['B5'].value),  # beta
            int(sheet['B6'].value),  # k
            int(sheet['G11'].value),  # s
            read_sample(sheet, 'C9', int(sheet['B6'].value))  # sample
        ),
        'beta': lambda: (
            float(sheet['B4'].value),  # alpha
            float(sheet['B5'].value),  # beta
            int(sheet['B6'].value),  # k
            int(sheet['G11'].value),  # s
            read_sample(sheet, 'C9', int(sheet['B6'].value))  # sample
        )
    }

    try:
        return params[dist_type]()
    except Exception as e:
        print(f"Ошибка чтения {dist_type} распределения: {e}")
        return (None,) * 5


def read_sample(sheet, start_cell: str, k: int) -> List[float]:
    """Чтение выборки из указанного диапазона"""
    sample = []
    start_row = int(''.join(filter(str.isdigit, start_cell)))
    start_col = ''.join(filter(str.isalpha, start_cell))

    for i in range(k):
        cell = f"{start_col}{start_row + i}"
        cell_value = sheet[cell].value
        if cell_value is not None:
            try:
                sample.append(float(cell_value))
            except ValueError:
                print(f"Ошибка: значение в ячейке {cell} не является числом.")
                return []
    return sample


def calculate_stats(sample: List[float]) -> dict:
    """Вычисление статистических характеристик"""
    return {
        'mean': np.mean(sample),
        'variance': np.var(sample),
        'std_dev': np.std(sample),
        'skewness': skew(sample),
        'kurtosis': kurtosis(sample)
    }


def print_stats(stats: dict, dist_name: str):
    """Печать статистических характеристик"""
    print(f"\nСтатистика {dist_name}:")
    print(f"Среднее: {stats['mean']:.4f}")
    print(f"Дисперсия: {stats['variance']:.4f}")
    print(f"СКО: {stats['std_dev']:.4f}")
    print(f"Ассиетрия: {stats['skewness']:.4f}")
    print(f"Эксцесс: {stats['kurtosis']:.4f}")


def plot_distribution(sample: List[float], bins: int, dist_name: str):
    """Построение графиков распределения"""
    plt.figure(figsize=(14, 10))

    # График значений
    plt.subplot(2, 2, 1)
    plt.plot(range(1, len(sample) + 1), sample, 'bo-')
    plt.title(f'{dist_name} - Значения')
    plt.xlabel('Индекс')
    plt.ylabel('Значение')

    # Гистограмма
    plt.subplot(2, 2, 2)
    plt.hist(sample, bins=bins, edgecolor='black')
    plt.title(f'{dist_name} - Гистограмма')
    plt.xlabel('Значение')
    plt.ylabel('Частота')

    # Функция распределения
    plt.subplot(2, 2, 3)
    hist, edges = np.histogram(sample, bins=bins, density=True)
    cdf = np.cumsum(hist) * np.diff(edges)
    plt.plot(edges[:-1], cdf, 'r-')
    plt.title(f'{dist_name} - Функция распределения')
    plt.xlabel('Значение')
    plt.ylabel('Вероятность')

    plt.tight_layout()
    plt.show()


def test_mode():
    """Тестовый режим - анализ данных из Excel"""
    wb = load_workbook()

    for dist_type in SHEET_NAMES:
        data = read_distribution_data(wb, dist_type)
        if None in data:
            continue

        if dist_type == 'uniform':
            a, b, k, s, sample = data
            print(f"\nРавномерное: a={a}, b={b}, k={k}")
            bins = np.linspace(a - (b - a), b + (b - a), s)
        else:
            param1, param2, k, s, sample = data
            print(f"\n{dist_type.capitalize()}: param1={param1}, param2={param2}, k={k}")
            bins = np.linspace(min(sample), max(sample), s)

        stats = calculate_stats(sample)
        print_stats(stats, dist_type)
        plot_distribution(sample, bins, dist_type.capitalize())

    wb.close()


def work_mode():
    """Интерактивный режим - генерация данных"""
    dist_types = {
        'uniform': {'params': ['a', 'b'], 'generator': np.random.uniform},
        'normal': {'params': ['mean', 'std'], 'generator': np.random.normal},
        'gamma': {'params': ['shape', 'scale'], 'generator': np.random.gamma},
        'beta': {'params': ['alpha', 'beta'], 'generator': np.random.beta}
    }

    for name, config in dist_types.items():
        print(f"\n{name.capitalize()} распределение:")
        params = [float(input(f"Введите {p}: ")) for p in config['params']]
        k = int(input("Объем выборки (k): "))
        s = int(input("Число карманов (s): "))

        sample = config['generator'](*params, size=k)

        if name == 'gamma':
            sample = config['generator'](params[0], 1 / params[1], size=k)
        elif name == 'beta':
            bins = np.linspace(0, 1, s)
        else:
            bins = np.linspace(min(sample), max(sample), s)

        stats = calculate_stats(sample)
        print_stats(stats, name)
        plot_distribution(sample, bins, name.capitalize())


def main():
    """Основная функция"""
    modes = {
        'test': test_mode,
        'work': work_mode,
        'exit': lambda: print("Выход")
    }

    while True:
        choice = input("\nРежим (test/work/exit): ").lower().strip()
        if choice in modes:
            if choice == 'exit':
                break
            modes[choice]()
        else:
            print("Неверный выбор!")


if __name__ == "__main__":
    main()